import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
try:
    from openpyxl.chart.label import DataLabelList
except ImportError:  # Older openpyxl builds
    DataLabelList = None
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from collections import Counter

FAIL_GRADES = {'F', 'FE', 'I', 'LP'}
GRADE_ORDER  = ['S', 'A+', 'A', 'B+', 'B', 'C+', 'C', 'D', 'P', 'F', 'FE', 'LP', 'I']
SEMS         = [f'S{i}' for i in range(1, 9)]

# ── Colour palettes ───────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SUB_FILL   = PatternFill("solid", fgColor="2E75B6")   # medium blue
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")   # light blue
FAIL_FILL  = PatternFill("solid", fgColor="FF0000")
WARN_FILL  = PatternFill("solid", fgColor="FFC000")
GOOD_FILL  = PatternFill("solid", fgColor="70AD47")
PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")

thin = Side(style='thin', color="AAAAAA")
medium = Side(style='medium', color="444444")
BORDER_THIN   = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
BORDER_MEDIUM = Border(left=medium, right=medium, top=medium, bottom=medium)


def _hdr(ws, text, row, col, bold=True, fg="FFFFFF", fill=None, align='center',
         size=11):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if fill:
        cell.fill = fill
    cell.border = BORDER_THIN
    return cell


def _val(ws, val, row, col, fill=None, bold=False, align='center', num_fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = BORDER_THIN
    if fill:
        cell.fill = fill
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _sgpa_fill(val):
    if not isinstance(val, (int, float)):
        return None
    if val >= 8.5:
        return GOOD_FILL
    if val >= 6.0:
        return PASS_FILL
    return FAIL_FILL


def _backlog_fill(val):
    if not isinstance(val, (int, float)):
        return None
    if val == 0:
        return GOOD_FILL
    if val <= 2:
        return WARN_FILL
    return FAIL_FILL


def _auto_col_widths(ws, min_w=8, max_w=40):
    for col in ws.columns:
        width = max(
            (len(str(cell.value)) if cell.value else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max_w, max(min_w, width + 2))


def _sem_key(sem):
    if isinstance(sem, str) and sem.startswith('S') and sem[1:].isdigit():
        return int(sem[1:])
    return 999


# ═════════════════════════════════════════════════════════════════════════════

def generate_excel_report(results_data, grades_data, students_data,
                           net_backlogs, output_filepath, selected_course_code=None):
    """
    results_data  – list of {register_no, semester, sgpa, backlogs}
    grades_data   – list of {register_no, semester, course_code, course_name, grade}
    students_data – list of {register_no, name}
    net_backlogs  – {register_no: {semester: count}}  from compute_net_backlogs()
    output_filepath – path to write .xlsx
    selected_course_code – optional course code for subject analysis sheet
    """
    wb = Workbook()
    wb.remove(wb.active)

    name_map = {s['register_no']: s['name'] for s in students_data}

    df_res    = pd.DataFrame(results_data)
    df_grades = pd.DataFrame(grades_data)

    sems_present = sorted(
        {r['semester'] for r in results_data},
        key=lambda s: int(s[1:])
    ) if results_data else []

    reg_nos = sorted({r['register_no'] for r in results_data})

    # ── 1. SGPA Matrix ────────────────────────────────────────────────────────
    ws = wb.create_sheet("SGPA Matrix")
    _write_matrix_sheet(
        ws, reg_nos, sems_present, df_res, name_map,
        value_key='sgpa',
        title='Semester-wise SGPA Matrix',
        fill_fn=_sgpa_fill,
        fmt='0.00'
    )

    # ── 2. Net Backlog Matrix ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Backlog Matrix")
    _write_backlog_sheet(ws2, reg_nos, sems_present, net_backlogs, name_map)

    # ── 3. Subject Analysis (for selected course only) ────────────────────────
    if selected_course_code and not df_grades.empty:
        sub = df_grades[df_grades['course_code'] == selected_course_code]
        if not sub.empty:
            course_name = sub['course_name'].iloc[0] or selected_course_code
            sem = sub['semester'].iloc[0]
            ws_c = wb.create_sheet(title="Subject Analysis")
            _write_course_sheet(ws_c, selected_course_code, course_name, sem, sub, name_map=name_map)

    if not wb.sheetnames:
        wb.create_sheet("No Data")

    wb.save(output_filepath)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _write_matrix_sheet(ws, reg_nos, sems, df_res, name_map,
                         value_key, title, fill_fn, fmt):
    ws.freeze_panes = 'D2'

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=3 + len(sems))
    cell = ws.cell(1, 1, title)
    cell.font      = Font(bold=True, size=14, color="FFFFFF")
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Column headers
    _hdr(ws, '#',            2, 1, fill=SUB_FILL)
    _hdr(ws, 'Register No',  2, 2, fill=SUB_FILL)
    _hdr(ws, 'Name',         2, 3, fill=SUB_FILL)
    for j, sem in enumerate(sems, start=4):
        _hdr(ws, sem, 2, j, fill=SUB_FILL)

    ws.row_dimensions[2].height = 20

    # Build lookup
    lookup = {}
    for _, r in df_res.iterrows():
        lookup.setdefault(r['register_no'], {})[r['semester']] = r[value_key]

    for i, rn in enumerate(reg_nos, start=1):
        r_row = i + 2
        fill  = ALT_FILL if i % 2 == 0 else None
        _val(ws, i,                  r_row, 1, fill=fill, align='center')
        _val(ws, rn,                 r_row, 2, fill=fill, align='left', bold=True)
        _val(ws, name_map.get(rn,''),r_row, 3, fill=fill, align='left')
        for j, sem in enumerate(sems, start=4):
            v = lookup.get(rn, {}).get(sem, '-')
            _val(ws, v, r_row, j, fill=fill_fn(v) or fill, num_fmt=fmt)

    _auto_col_widths(ws)
    ws.column_dimensions['C'].width = 28


def _write_backlog_sheet(ws, reg_nos, sems, net_backlogs, name_map):
    ws.freeze_panes = 'D2'

    # Title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=3 + len(sems))
    cell = ws.cell(1, 1, 'Net Cumulative Backlog Matrix')
    cell.font      = Font(bold=True, size=14, color="FFFFFF")
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    _hdr(ws, '#',            2, 1, fill=SUB_FILL)
    _hdr(ws, 'Register No',  2, 2, fill=SUB_FILL)
    _hdr(ws, 'Name',         2, 3, fill=SUB_FILL)
    for j, sem in enumerate(sems, start=4):
        _hdr(ws, sem, 2, j, fill=SUB_FILL)

    ws.row_dimensions[2].height = 20

    for i, rn in enumerate(reg_nos, start=1):
        r_row = i + 2
        base  = ALT_FILL if i % 2 == 0 else None
        _val(ws, i,                   r_row, 1, fill=base, align='center')
        _val(ws, rn,                  r_row, 2, fill=base, align='left', bold=True)
        _val(ws, name_map.get(rn,''), r_row, 3, fill=base, align='left')
        for j, sem in enumerate(sems, start=4):
            v = net_backlogs.get(rn, {}).get(sem, '-')
            cell_fill = _backlog_fill(v) if isinstance(v, int) else base
            cell = _val(ws, v, r_row, j, fill=cell_fill, align='center')
            if v == 0:
                cell.value = '✓ 0'

    _auto_col_widths(ws)
    ws.column_dimensions['C'].width = 28


def _write_course_sheet(ws, code, course_name, sem, sub_df, name_map=None):
    GRADE_POINTS = {'S': 10, 'A+': 9, 'A': 8.5, 'B+': 8, 'B': 7.5,
                    'C+': 7, 'C': 6.5, 'D': 6, 'P': 5, 'F': 0, 'FE': 0, 'LP': 0, 'I': 0}
    QUALITY_GRADES = {'S', 'A+', 'A', 'B+'}

    # Header
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value     = f"{code} – {course_name}"
    cell.font      = Font(bold=True, size=13, color="FFFFFF")
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws['A2'] = f'Semester: {sem}    |    Total Students: {len(sub_df)}'
    ws['A2'].font      = Font(italic=True, size=10, color="444444")
    ws['A2'].alignment = Alignment(horizontal='left')
    ws.merge_cells('A2:H2')

    # Grade distribution table
    _hdr(ws, 'Grade',  4, 1, fill=SUB_FILL)
    _hdr(ws, 'Count',  4, 2, fill=SUB_FILL)
    _hdr(ws, '%',      4, 3, fill=SUB_FILL)
    _hdr(ws, 'Status', 4, 4, fill=SUB_FILL)

    counts = Counter(sub_df['grade'])
    total  = len(sub_df)
    data_row = 5

    for g in GRADE_ORDER:
        if g not in counts:
            continue
        cnt  = counts[g]
        pct  = round(cnt / total * 100, 1) if total else 0
        is_f = g in FAIL_GRADES
        fill = FAIL_FILL if is_f else PASS_FILL
        _val(ws, g,          data_row, 1, fill=fill, bold=True, align='center')
        _val(ws, cnt,        data_row, 2, fill=fill, align='center')
        _val(ws, f'{pct}%',  data_row, 3, fill=fill, align='center')
        _val(ws, 'Fail' if is_f else 'Pass', data_row, 4,
             fill=fill, align='center')
        data_row += 1

    # Summary row
    pass_cnt = sum(v for k, v in counts.items() if k not in FAIL_GRADES)
    fail_cnt = total - pass_cnt
    pass_pct = round(pass_cnt / total * 100, 1) if total else 0
    _val(ws, 'TOTAL',    data_row, 1, fill=SUB_FILL, bold=True, align='center')
    _val(ws, total,      data_row, 2, fill=SUB_FILL, bold=True, align='center')
    _val(ws, '100%',     data_row, 3, fill=SUB_FILL, bold=True, align='center')
    _val(ws, f'Pass: {pass_cnt} | Fail: {fail_cnt}', data_row, 4,
         fill=SUB_FILL, bold=True, align='center')
    ws.merge_cells(start_row=data_row, start_column=4,
                   end_row=data_row,   end_column=6)

    # ── Summary Statistics ────────────────────────────────────────────────────
    stats_row = data_row + 2
    STATS_FILL = PatternFill("solid", fgColor="FFF2CC")

    # Avg Grade Point
    gp_sum = sum(GRADE_POINTS.get(row['grade'], 0) for _, row in sub_df.iterrows())
    avg_gp = round(gp_sum / total, 2) if total else 0

    # Quality Index (% ≥ B+)
    quality_cnt = sum(counts.get(g, 0) for g in QUALITY_GRADES)
    quality_idx = round(quality_cnt / total * 100, 1) if total else 0

    ws.merge_cells(start_row=stats_row, start_column=1,
                   end_row=stats_row,   end_column=6)
    _hdr(ws, 'SUMMARY STATISTICS', stats_row, 1, fill=HDR_FILL,
         fg="FFFFFF", align='center', size=11)

    labels_vals = [
        ('Avg Grade Point', avg_gp),
        ('Pass Percentage', f'{pass_pct}%'),
        ('Quality Index (≥B+)', f'{quality_idx}%'),
    ]
    for k, (lbl, val) in enumerate(labels_vals):
        r = stats_row + 1 + k
        _val(ws, lbl, r, 1, fill=STATS_FILL, bold=True, align='left')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _val(ws, val, r, 3, fill=STATS_FILL, bold=True, align='center')

    chart_anchor_row = stats_row + 1 + len(labels_vals) + 1

    # ── Pie Chart ─────────────────────────────────────────────────────────────
    if len(counts) > 0:
        pie = PieChart()
        data_ref   = Reference(ws, min_col=2, min_row=4,
                               max_row=3 + len(counts))
        labels_ref = Reference(ws, min_col=1, min_row=5,
                               max_row=4 + len(counts))
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        pie.title  = f"Grade Distribution: {code}"
        pie.style  = 10
        pie.width  = 14
        pie.height = 12
        ws.add_chart(pie, f"F4")

    # ── Bar Chart ─────────────────────────────────────────────────────────────
    if len(counts) > 0:
        bar = BarChart()
        bar_data_ref   = Reference(ws, min_col=2, min_row=4,
                                   max_row=3 + len(counts))
        bar_labels_ref = Reference(ws, min_col=1, min_row=5,
                                   max_row=4 + len(counts))
        bar.add_data(bar_data_ref, titles_from_data=True)
        bar.set_categories(bar_labels_ref)
        bar.title  = f"Grade Counts: {code}"
        bar.style  = 10
        bar.type   = "col"
        bar.width  = 16
        bar.height = 12
        bar.y_axis.title = "No. of Students"
        bar.x_axis.title = "Grade"
        ws.add_chart(bar, f"F{chart_anchor_row}")

    # ── Student-wise grade list ───────────────────────────────────────────────
    list_start = chart_anchor_row + 1
    _hdr(ws, 'Register No', list_start, 1, fill=SUB_FILL)
    _hdr(ws, 'Name',        list_start, 2, fill=SUB_FILL)
    _hdr(ws, 'Grade',       list_start, 3, fill=SUB_FILL)

    for idx, (_, row) in enumerate(sub_df.iterrows()):
        r  = list_start + 1 + idx
        is_f = row['grade'] in FAIL_GRADES
        fill = PatternFill("solid", fgColor="FFE0E0") if is_f else None
        _val(ws, row['register_no'], r, 1, fill=fill, align='left')
        student_name = ''
        if name_map:
            student_name = name_map.get(row['register_no'], '')
        _val(ws, student_name,       r, 2, fill=fill, align='left')
        _val(ws, row['grade'],       r, 3, fill=fill, align='center', bold=is_f)

    _auto_col_widths(ws)
    ws.column_dimensions['B'].width = 28


def _safe_sheetname(name):
    """Truncate and strip chars that Excel disallows in sheet names."""
    invalid = r'/\?*[]:'
    for ch in invalid:
        name = name.replace(ch, '')
    return name[:31]


def generate_subject_excel_report(grades_data, students_data, output_filepath,
                                   course_code, semester):
    """Generate an Excel report for a single subject (course_code + semester)."""
    wb = Workbook()
    wb.remove(wb.active)

    name_map = {s['register_no']: s['name'] for s in students_data}
    df_grades = pd.DataFrame(grades_data)

    if not df_grades.empty:
        sub = df_grades[(df_grades['course_code'] == course_code) &
                        (df_grades['semester'] == semester)]
        if not sub.empty:
            course_name = sub['course_name'].iloc[0] or course_code
            ws = wb.create_sheet(title=_safe_sheetname(course_code))
            _write_course_sheet(ws, course_code, course_name, semester, sub,
                                name_map=name_map)

    if not wb.sheetnames:
        wb.create_sheet("No Data")

    wb.save(output_filepath)


def generate_all_subjects_excel_report(grades_data, students_data, courses_data,
                                        semester, output_filepath,
                                        results_data=None, net_backlogs=None):
    """
    Generate an Excel report with a sheet for EACH subject in the semester,
    plus an overview summary sheet.

    Parameters:
        grades_data   – list of {register_no, semester, course_code, course_name, grade}
        students_data – list of {register_no, name}
        courses_data  – list of {code, name}
        semester      – e.g. 'S5'
        output_filepath – path to write .xlsx
    """
    wb = Workbook()
    wb.remove(wb.active)

    name_map = {s['register_no']: s['name'] for s in students_data}
    df_grades = pd.DataFrame(grades_data)

    if df_grades.empty:
        wb.create_sheet("No Data")
        wb.save(output_filepath)
        return

    QUALITY_GRADES = {'S', 'A+', 'A', 'B+'}
    GRADE_PTS = {'S': 10, 'A+': 9, 'A': 8.5, 'B+': 8, 'B': 7.5,
                 'C+': 7, 'C': 6.5, 'D': 6, 'P': 5,
                 'F': 0, 'FE': 0, 'LP': 0, 'I': 0, 'Absent': 0}

    summary_rows = []   # for the overview sheet

    for course in sorted(courses_data, key=lambda c: c['code']):
        code = course['code']
        c_name = course['name'] or code

        sub = df_grades[(df_grades['course_code'] == code) &
                        (df_grades['semester'] == semester)]
        if sub.empty:
            continue

        # Write per-subject sheet
        ws = wb.create_sheet(title=_safe_sheetname(code))
        _write_course_sheet(ws, code, c_name, semester, sub, name_map=name_map)

        # Build summary row
        counts = Counter(sub['grade'])
        total = len(sub)
        pass_cnt = sum(v for k, v in counts.items() if k not in FAIL_GRADES)
        fail_cnt = total - pass_cnt
        pass_pct = round(pass_cnt / total * 100, 1) if total else 0
        gp_sum = sum(GRADE_PTS.get(row['grade'], 0) for _, row in sub.iterrows())
        avg_gp = round(gp_sum / total, 2) if total else 0
        quality_cnt = sum(counts.get(g, 0) for g in QUALITY_GRADES)
        quality_idx = round(quality_cnt / total * 100, 1) if total else 0

        summary_rows.append({
            'code': code,
            'name': c_name,
            'total': total,
            'pass_cnt': pass_cnt,
            'fail_cnt': fail_cnt,
            'pass_pct': pass_pct,
            'avg_gp': avg_gp,
            'quality_idx': quality_idx,
        })

    # ── Overview Sheet ────────────────────────────────────────────────────────
    if summary_rows:
        ws_ov = wb.create_sheet(title="Overview", index=0)  # first sheet
        ws_ov.freeze_panes = 'A3'

        ws_ov.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=8)
        cell = ws_ov.cell(1, 1, f'All Subjects – Semester {semester}')
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_ov.row_dimensions[1].height = 26

        headers = ['Course Code', 'Course Name', 'Total', 'Passed', 'Failed',
                   'Pass %', 'Avg GP', 'Quality Index (≥B+)']
        for j, h in enumerate(headers, 1):
            _hdr(ws_ov, h, 2, j, fill=SUB_FILL)
        ws_ov.row_dimensions[2].height = 20

        for i, r in enumerate(summary_rows, start=1):
            row_num = i + 2
            base = ALT_FILL if i % 2 == 0 else None
            _val(ws_ov, r['code'],        row_num, 1, fill=base, bold=True, align='left')
            _val(ws_ov, r['name'],        row_num, 2, fill=base, align='left')
            _val(ws_ov, r['total'],       row_num, 3, fill=base)
            _val(ws_ov, r['pass_cnt'],    row_num, 4,
                 fill=GOOD_FILL if r['pass_pct'] >= 80 else base)
            _val(ws_ov, r['fail_cnt'],    row_num, 5,
                 fill=FAIL_FILL if r['fail_cnt'] > r['total'] * 0.3 else base)
            _val(ws_ov, f"{r['pass_pct']}%", row_num, 6, fill=base, bold=True)
            _val(ws_ov, r['avg_gp'],      row_num, 7, fill=base, num_fmt='0.00')
            _val(ws_ov, f"{r['quality_idx']}%", row_num, 8, fill=base)

        _auto_col_widths(ws_ov)
        ws_ov.column_dimensions['B'].width = 36

        # ── Summary Bar Chart (screenshot-style: one PASSED category) ────
        if len(summary_rows) > 0:
            chart_tbl_col = 10  # J
            chart_tbl_row = len(summary_rows) + 5

            # Build helper table:
            # Row 1 -> subject codes (series names)
            # Row 2 -> passed counts with a single category label "PASSED"
            ws_ov.cell(row=chart_tbl_row, column=chart_tbl_col, value='')
            for idx, r in enumerate(summary_rows, start=1):
                ws_ov.cell(row=chart_tbl_row, column=chart_tbl_col + idx, value=r['code'])
                ws_ov.cell(row=chart_tbl_row + 1, column=chart_tbl_col + idx, value=r['pass_cnt'])
            ws_ov.cell(row=chart_tbl_row + 1, column=chart_tbl_col, value='PASSED')

            bar = BarChart()
            bar.title = "Chart Title"
            bar.type = "bar"
            bar.style = 10
            bar.width = 18
            bar.height = 10
            bar.x_axis.title = "Students"
            bar.y_axis.title = ""

            data_ref = Reference(
                ws_ov,
                min_col=chart_tbl_col + 1,
                min_row=chart_tbl_row,
                max_col=chart_tbl_col + len(summary_rows),
                max_row=chart_tbl_row + 1,
            )
            cats_ref = Reference(
                ws_ov,
                min_col=chart_tbl_col,
                min_row=chart_tbl_row + 1,
                max_row=chart_tbl_row + 1,
            )
            bar.add_data(data_ref, titles_from_data=True)
            bar.set_categories(cats_ref)

            # Keep legend at bottom and hide dense labels to prevent squishing.
            if bar.legend is not None:
                bar.legend.position = "b"

            ws_ov.add_chart(bar, f"A{len(summary_rows) + 6}")

    # ── SGPA and Backlog Matrices ───────────────────────────────────────────
    results_data = results_data or []
    net_backlogs = net_backlogs or {}

    if results_data:
        df_res = pd.DataFrame(results_data)
        sems_present = sorted(
            {
                r.get('semester')
                for r in results_data
                if isinstance(r.get('semester'), str)
                and r.get('semester', '').startswith('S')
                and r.get('semester', '')[1:].isdigit()
            },
            key=_sem_key,
        )
        reg_nos = sorted({s['register_no'] for s in students_data if s.get('register_no')})

        if not sems_present:
            sems_present = [semester]
        if not reg_nos:
            reg_nos = sorted({r.get('register_no') for r in results_data if r.get('register_no')})

        ws_sgpa = wb.create_sheet("SGPA Matrix", index=1)
        _write_matrix_sheet(
            ws_sgpa,
            reg_nos,
            sems_present,
            df_res,
            name_map,
            value_key='sgpa',
            title='Semester-wise SGPA Matrix',
            fill_fn=_sgpa_fill,
            fmt='0.00'
        )

        ws_backlog = wb.create_sheet("Backlog Matrix", index=2)
        _write_backlog_sheet(ws_backlog, reg_nos, sems_present, net_backlogs, name_map)

    if not wb.sheetnames:
        wb.create_sheet("No Data")

    wb.save(output_filepath)

